"""
Excel workbook builder for ALLL WPS Designer.

Builds a professional .xlsx workbook from a serialised ProjectDraft dict.

Sheets
------
 1. Inputs Summary          — Parameter | Display value | SI value | Unit
 2. Hydraulics Breakdown    — per-segment rows (suction then discharge) + totals
 3. System Curve            — Q/H table; chart: HQ pump curves + system + duty pts
 4. Pump Curves             — Q/H/η/P/NPSHr table; chart: η vs Q
 5. Operating Points        — duty table; charts: P vs Q and NPSHa vs NPSHr
 6. Wet Well                — geometry KPIs, volume curve, cycle results
 7. Engineering Checks      — velocity, Re, NPSH, cycling, surge checks
 8. Surge Quick (Mode A)    — suction section + discharge section
 9. Surge MOC Time Histories— suction section + discharge section (≤ 1000 rows each)
10. Surge Envelope vs Dist  — suction section + discharge section; chart per pipeline
11. Protection Comparisons  — baseline + device what-if runs

Charts: Sheets 3, 4, 5, 9, 10 each have at least one embedded chart (6 total).
"""

from __future__ import annotations

import io
import math
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Number-format policy (per spec: 2 dp heads, 0 dp flows, 1 dp efficiency)
# ---------------------------------------------------------------------------

FMT_FLOW  = "#,##0"        # m³/h, L/s — 0 decimal places
FMT_HEAD  = "#,##0.00"     # m      — 2 decimal places
FMT_EFF   = "#,##0.0"      # %      — 1 decimal place
FMT_POW   = "#,##0.00"     # kW     — 2 decimal places
FMT_NPSH  = "#,##0.00"     # m      — 2 decimal places
FMT_PRES  = "#,##0.0"      # kPa    — 1 decimal place
FMT_VEL   = "#,##0.000"    # m/s    — 3 decimal places
FMT_DIST  = "#,##0.0"      # m dist — 1 decimal place

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

G_ACC   = 9.81          # m/s²
MU_WATER = 0.001        # dynamic viscosity at ~20 °C [Pa·s]
RHO_WATER = 1000.0      # [kg/m³]

ROUGHNESS_M: dict[str, float] = {
    "pvc":      0.0000015,
    "hdpe":     0.0000015,
    "upvc":     0.0000015,
    "dicl":     0.00025,
    "ductile_iron": 0.00025,
    "steel":    0.000046,
    "ac":       0.00003,
    "asbestos_cement": 0.00003,
    "concrete": 0.0003,
    "grp":      0.00003,
    "frp":      0.00003,
    "default":  0.0001,
}

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

NAVY      = "1F3864"
BLUE_MID  = "2E5B8C"
BLUE_LITE = "D0E4F5"
TEAL_HDR  = "1A5276"
WHITE     = "FFFFFF"
ALT_ROW   = "F2F5F8"
GREY_SECT = "4A4A4A"

SEV_COLORS = {"OK": "27AE60", "WARNING": "F39C12", "CRITICAL": "E74C3C"}

# ---------------------------------------------------------------------------
# Unit conversion helpers
# ---------------------------------------------------------------------------

def _us(unit_system: str, m_val: float | None, kind: str) -> tuple[float | None, str]:
    """
    Convert SI value to display unit. Returns (display_value, unit_label).
    kind: 'flow' | 'head' | 'pressure' | 'velocity' | 'length' | 'volume' | 'power' | 'temp'
    """
    if m_val is None:
        return None, "—"
    if unit_system != "US":
        labels = {"flow": "m³/h", "head": "m", "pressure": "kPa",
                  "velocity": "m/s", "length": "m", "volume": "m³",
                  "power": "kW", "temp": "°C", "area": "m²"}
        return m_val, labels.get(kind, "—")
    conversions = {
        "flow":     (m_val * 4.40287, "gpm"),
        "head":     (m_val * 3.28084, "ft"),
        "pressure": (m_val * 0.14504, "psi"),
        "velocity": (m_val * 3.28084, "ft/s"),
        "length":   (m_val * 3.28084, "ft"),
        "volume":   (m_val * 264.172, "gal"),
        "power":    (m_val * 1.34102, "hp"),
        "temp":     (m_val * 9 / 5 + 32, "°F"),
        "area":     (m_val * 10.7639, "ft²"),
    }
    return conversions.get(kind, (m_val, "—"))


# ---------------------------------------------------------------------------
# Friction factor — Swamee-Jain approximation
# ---------------------------------------------------------------------------

def _friction_factor_sw(Re: float, eps_D: float) -> float:
    """Swamee-Jain explicit approximation of Darcy-Weisbach f."""
    if Re < 2300:
        return 64.0 / max(Re, 1.0)
    denom = math.log10(eps_D / 3.7 + 5.74 / max(Re, 1.0) ** 0.9)
    return 0.25 / denom ** 2


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _fill(hex_col: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_col)


def _border() -> Border:
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)


def _hdr(ws, row: int, col: int, val: str,
         bg: str = BLUE_MID, fg: str = WHITE, size: int = 10,
         merge_to: int | None = None) -> None:
    c = ws.cell(row=row, column=col, value=val)
    c.fill      = _fill(bg)
    c.font      = Font(bold=True, size=size, color=fg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _border()
    if merge_to:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=merge_to
        )


def _dat(ws, row: int, col: int, val: Any,
         fmt: str | None = None, bold: bool = False,
         align: str = "left", alt: bool = False,
         merge_to: int | None = None) -> None:
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _border()
    if alt:
        c.fill = _fill(ALT_ROW)
    if fmt and val is not None:
        c.number_format = fmt
    if merge_to:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=merge_to
        )


def _title_banner(ws, title: str, subtitle: str = "") -> None:
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value     = title
    c.fill      = _fill(NAVY)
    c.font      = Font(bold=True, size=14, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells("A2:L2")
        s = ws["A2"]
        s.value     = subtitle
        s.fill      = _fill(BLUE_MID)
        s.font      = Font(italic=True, size=10, color=WHITE)
        s.alignment = Alignment(horizontal="left", vertical="center")


def _meta_rows(ws, meta: dict, row: int) -> int:
    """Write project meta fields; return next available row."""
    fields = [
        ("Project",  meta.get("name", "")),
        ("Client",   meta.get("client", "")),
        ("Job No.",  meta.get("job_number", "")),
        ("Date",     meta.get("date", "")),
        ("Engineer", meta.get("engineer", "")),
    ]
    for lbl, val in fields:
        ws.cell(row=row, column=1, value=lbl).font  = Font(bold=True, size=9, color="444444")
        ws.cell(row=row, column=2, value=val).font  = Font(size=9)
        row += 1
    return row + 1


def _no_data(ws, row: int, msg: str = "Not computed — run analyses in WPS Designer first.") -> None:
    ws.merge_cells(f"A{row}:L{row}")
    c = ws[f"A{row}"]
    c.value     = f"N/A — {msg}"
    c.font      = Font(italic=True, color="888888", size=10)
    c.alignment = Alignment(horizontal="center")


def _section_hdr(ws, row: int, label: str, ncols: int = 10) -> None:
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=ncols
    )
    c = ws.cell(row=row, column=1, value=label)
    c.fill      = _fill(TEAL_HDR)
    c.font      = Font(bold=True, size=11, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18


def _col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _nearest(mapping: dict, q: float) -> float | None:
    """Return value from mapping whose key is nearest to q, or None if empty."""
    if not mapping:
        return None
    key = min(mapping.keys(), key=lambda k: abs(k - q))
    return mapping[key]


def _subsample(rows: list[dict], max_rows: int = 1000) -> list[dict]:
    """Sub-sample a list of dicts to at most max_rows entries."""
    if len(rows) <= max_rows:
        return rows
    step = math.ceil(len(rows) / max_rows)
    return rows[::step]


# ---------------------------------------------------------------------------
# Sheet 1 — Inputs Summary
# ---------------------------------------------------------------------------

def _sh_inputs(wb: Workbook, draft: dict) -> None:
    ws   = wb.create_sheet("Inputs Summary")
    meta = draft.get("meta", {}) or {}
    us   = draft.get("unitSystem", "SI")

    _title_banner(ws, "WPS Designer — Inputs Summary",
                  "Design inputs and pipeline configuration")
    row = 3
    row = _meta_rows(ws, meta, row)

    Q_m3h = draft.get("designFlow_m3h", 0.0) or 0.0
    up    = draft.get("upstreamNode",   {}) or {}
    dn    = draft.get("downstreamNode", {}) or {}

    # Column headers
    _hdr(ws, row, 1, "Parameter");    _hdr(ws, row, 2, "Display value")
    _hdr(ws, row, 3, "SI value");     _hdr(ws, row, 4, "Unit (display)");
    _hdr(ws, row, 5, "SI unit")
    row += 1

    static_m = dn.get("elevation_m", 0.0) - up.get("elevation_m", 0.0)

    def add_row(param: str, si_val: float | str | None, kind: str = "",
                si_unit: str = "—", alt: bool = False) -> None:
        nonlocal row
        if kind and isinstance(si_val, (int, float)):
            disp_val, disp_unit = _us(us, float(si_val), kind)
        else:
            disp_val, disp_unit = si_val, si_unit
        _dat(ws, row, 1, param,     bold=True, alt=alt)
        _dat(ws, row, 2, disp_val,  fmt="#,##0.000" if isinstance(disp_val, float) else None,
             align="right", alt=alt)
        _dat(ws, row, 3, si_val,    fmt="#,##0.000" if isinstance(si_val, float) else None,
             align="right", alt=alt)
        _dat(ws, row, 4, disp_unit, align="center", alt=alt)
        _dat(ws, row, 5, si_unit,   align="center", alt=alt)
        row += 1

    add_row("Unit system",            us,                                   si_unit="—")
    add_row("Design flow Q",          Q_m3h,                               "flow",     "m³/h")
    add_row("Design flow Q (L/s)",    round(Q_m3h / 3.6, 4),              "flow",     "m³/h", alt=True)
    add_row("Upstream elevation",     up.get("elevation_m", 0.0),          "head",     "m")
    add_row("Upstream pressure",      up.get("pressure_kPa", 0.0),         "pressure", "kPa",  alt=True)
    add_row("Downstream elevation",   dn.get("elevation_m", 0.0),          "head",     "m")
    add_row("Downstream pressure",    dn.get("pressure_kPa", 0.0),         "pressure", "kPa",  alt=True)
    add_row("Static head",            round(static_m, 3),                  "head",     "m")

    # Pipeline segments tables
    row += 1
    for pipe_key, title in [("suction", "Suction Pipeline Segments"),
                             ("discharge", "Discharge Pipeline Segments")]:
        _section_hdr(ws, row, title)
        row += 1
        segs = (draft.get(pipe_key) or {}).get("segments", [])
        k_sum = (draft.get(pipe_key) or {}).get("accessories_K_sum", 0.0)
        col_hdrs = ["Seg #", "Material", "DN (mm)", "L (m)", "L display", "Cumul. (m)"]
        for ci, h in enumerate(col_hdrs, 1):
            _hdr(ws, row, ci, h, BLUE_LITE, NAVY, 9)
        row += 1
        cum = 0.0
        for j, seg in enumerate(segs):
            l_m  = seg.get("length_m",    0.0)
            d_mm = seg.get("diameter_mm", 0.0)
            cum += l_m
            alt  = (j % 2 == 1)
            disp_l, disp_l_unit = _us(us, l_m, "length")
            _dat(ws, row, 1, j + 1,                    align="center", alt=alt)
            _dat(ws, row, 2, seg.get("material", ""),  alt=alt)
            _dat(ws, row, 3, d_mm,                     fmt="0.0", align="right", alt=alt)
            _dat(ws, row, 4, l_m,                      fmt="#,##0.00", align="right", alt=alt)
            _dat(ws, row, 5, f"{disp_l:.2f} {disp_l_unit}", alt=alt)
            _dat(ws, row, 6, round(cum, 2),             fmt="#,##0.00", align="right", alt=alt)
            row += 1
        _dat(ws, row, 1, "Totals",             bold=True, alt=True)
        _dat(ws, row, 2, f"{len(segs)} seg",   alt=True)
        _dat(ws, row, 3, "ΣK = " + str(round(k_sum, 4)), alt=True)
        _dat(ws, row, 4, round(cum, 2),         fmt="#,##0.00", align="right", alt=True)
        row += 2

    _col_widths(ws, [30, 16, 14, 12, 16, 14])


# ---------------------------------------------------------------------------
# Sheet 2 — Hydraulics Breakdown (per-segment)
# ---------------------------------------------------------------------------

def _seg_darcy(Q_m3h: float, seg: dict) -> dict:
    """Compute per-segment Darcy-Weisbach hydraulics from geometry + Q."""
    D_m   = seg.get("diameter_mm", 150.0) / 1000.0
    L_m   = seg.get("length_m",    0.0)
    mat   = seg.get("material",    "default").lower()
    eps   = ROUGHNESS_M.get(mat, ROUGHNESS_M["default"])
    A     = math.pi * D_m ** 2 / 4.0
    Q_m3s = Q_m3h / 3600.0
    v     = Q_m3s / max(A, 1e-9)
    Re    = RHO_WATER * v * D_m / MU_WATER
    eps_D = eps / max(D_m, 1e-9)
    f     = _friction_factor_sw(Re, eps_D)
    Hf    = f * L_m / D_m * v ** 2 / (2 * G_ACC)
    return {"v": v, "Re": Re, "f": f, "Hf": Hf, "D_m": D_m, "L_m": L_m}


def _sh_hydraulics(wb: Workbook, draft: dict) -> None:
    ws   = wb.create_sheet("Hydraulics Breakdown")
    meta = draft.get("meta", {}) or {}

    _title_banner(ws, "Hydraulic Calculation — Per-Segment Breakdown",
                  "Darcy-Weisbach friction loss per segment + aggregate results")
    row = 3
    row = _meta_rows(ws, meta, row)

    Q = draft.get("designFlow_m3h", 0.0) or 0.0
    r = draft.get("hydraulicsResult") or {}
    us = draft.get("unitSystem", "SI")

    col_hdrs = ["Seg #", "Material", "DN (mm)", "L (m)", "v (m/s)",
                "Re", "f (D-W)", "Hf (m)", "Hm (m)", "Hf display", "Cumul. Hf (m)"]
    for ci, h in enumerate(col_hdrs, 1):
        _hdr(ws, row, ci, h, BLUE_MID)
    row += 1
    ws.freeze_panes = f"A{row}"

    cum_hf = 0.0
    grand_hf = 0.0
    for pipe_key, pipe_lbl in [("suction", "SUCTION"), ("discharge", "DISCHARGE")]:
        _section_hdr(ws, row, f"— {pipe_lbl} pipeline —", 11)
        row += 1
        pipe     = draft.get(pipe_key) or {}
        segs     = pipe.get("segments", [])
        K_total  = pipe.get("accessories_K_sum", 0.0) or 0.0
        n_segs   = len(segs) if segs else 1
        seg_hf_total = 0.0
        for j, seg in enumerate(segs):
            s   = _seg_darcy(Q, seg)
            alt = (j % 2 == 1)
            # Distribute pipeline-level minor loss K equally across segments
            K_seg = K_total / n_segs
            Hm_seg = K_seg * s["v"] ** 2 / (2 * G_ACC) if s["v"] else 0.0
            disp_hf, disp_hf_unit = _us(us, s["Hf"], "head")
            _dat(ws, row,  1, j + 1,                    align="center", alt=alt)
            _dat(ws, row,  2, seg.get("material", ""),   alt=alt)
            _dat(ws, row,  3, seg.get("diameter_mm", 0), fmt="0.0", align="right", alt=alt)
            _dat(ws, row,  4, s["L_m"],                  fmt=FMT_DIST, align="right", alt=alt)
            _dat(ws, row,  5, round(s["v"],   4),        fmt=FMT_VEL,  align="right", alt=alt)
            _dat(ws, row,  6, int(s["Re"]),               fmt=FMT_FLOW, align="right", alt=alt)
            _dat(ws, row,  7, round(s["f"],   5),        fmt="0.00000", align="right", alt=alt)
            _dat(ws, row,  8, round(s["Hf"],  3),        fmt=FMT_HEAD, align="right", alt=alt)
            _dat(ws, row,  9, round(Hm_seg,   3),        fmt=FMT_HEAD, align="right", alt=alt)
            _dat(ws, row, 10, f"{disp_hf:.3f} {disp_hf_unit}", alt=alt)
            cum_hf += s["Hf"]
            seg_hf_total += s["Hf"]
            _dat(ws, row, 11, round(cum_hf, 3),          fmt=FMT_HEAD, align="right", alt=alt)
            row += 1
        grand_hf += seg_hf_total
        # Subtotal row
        _dat(ws, row, 1, f"{pipe_lbl} subtotal", bold=True, alt=True)
        _dat(ws, row, 8, round(seg_hf_total, 3), fmt=FMT_HEAD, align="right", bold=True, alt=True)
        row += 1

    row += 1
    # Totals row from aggregate result if available
    _section_hdr(ws, row, "— AGGREGATE TOTALS (from compute engine) —", 10)
    row += 1
    agg_rows = [
        ("Friction head Hf (engine)",  r.get("friction_head_m"), "head"),
        ("Minor head Hm (engine)",     r.get("minor_head_m"),    "head"),
        ("Static head (engine)",       r.get("static_head_m"),   "head"),
        ("TDH (engine)",               r.get("tdh_m"),           "head"),
        ("Velocity (engine)",          r.get("velocity_ms"),     "velocity"),
        ("Reynolds number (engine)",   r.get("reynolds_number"), ""),
        ("Friction factor f (engine)", r.get("friction_factor"), ""),
    ]
    for i, (lbl, val, kind) in enumerate(agg_rows):
        alt = (i % 2 == 1)
        if kind and isinstance(val, (int, float)):
            dv, du = _us(us, float(val), kind)
        else:
            dv, du = val, "—"
        _dat(ws, row, 1, lbl,  bold=True, alt=alt)
        _dat(ws, row, 2, val,  fmt="#,##0.0000" if isinstance(val, float) else None,
             align="right", alt=alt)
        _dat(ws, row, 3, "SI", alt=alt)
        _dat(ws, row, 4, dv,   fmt="#,##0.0000" if isinstance(dv, float) else None,
             align="right", alt=alt)
        _dat(ws, row, 5, du,   align="center", alt=alt)
        row += 1

    if not r:
        _no_data(ws, row)

    _col_widths(ws, [28, 14, 10, 10, 12, 12, 10, 12, 12, 18, 16])


# ---------------------------------------------------------------------------
# Sheet 3 — System Curve  (chart: HQ + system + duty pts)
# ---------------------------------------------------------------------------

def _sh_system_curve(wb: Workbook, draft: dict) -> None:
    ws   = wb.create_sheet("System Curve")
    meta = draft.get("meta", {}) or {}

    _title_banner(ws, "System Curve — H vs Q",
                  "Tabulated system resistance + pump H-Q overlay (rated speed + VFD speeds)")
    row = 3
    row = _meta_rows(ws, meta, row)

    r  = draft.get("hydraulicsResult") or {}
    pr = draft.get("pumpResult")        or {}
    sys_pts     = r.get("system_curve", [])
    hq_pts      = pr.get("hq_curve", [])
    ops         = pr.get("operating_points", [])
    speed_curves = pr.get("speed_curves", [])

    if not sys_pts and not hq_pts:
        _no_data(ws, row)
        return

    data_start = row
    hdrs = ["Q (m³/h)", "H_sys (m)", "H_pump 100% (m)"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, row, ci, h, BLUE_MID)
    row += 1
    ws.freeze_panes = f"A{row}"

    # Merge system curve and pump HQ by Q index
    n = max(len(sys_pts), len(hq_pts))
    for i in range(n):
        sys_q = sys_pts[i].get("Q_m3h", "") if i < len(sys_pts) else ""
        sys_h = sys_pts[i].get("H_m",   "") if i < len(sys_pts) else ""
        hq_h  = hq_pts[i].get("value",  "") if i < len(hq_pts)  else ""
        alt   = (i % 2 == 1)
        _dat(ws, row, 1,
             sys_q if sys_q != "" else (hq_pts[i].get("Q_m3h", "") if i < len(hq_pts) else ""),
             fmt=FMT_FLOW, align="right", alt=alt)
        _dat(ws, row, 2, sys_h, fmt=FMT_HEAD, align="right", alt=alt)
        _dat(ws, row, 3, hq_h,  fmt=FMT_HEAD, align="right", alt=alt)
        row += 1

    data_end = row - 1

    # ── write speed-curve auxiliary columns for chart overlays ──────────
    aux_col = 5          # start after a blank col 4 gap
    spd_series: list[tuple[int, int, int, int, str]] = []
    for sc in speed_curves:
        spd  = sc.get("speed_pct", 100)
        pts  = sc.get("hq_pts", [])
        ws.cell(row=data_start, column=aux_col,     value=f"{spd}% speed Q (m³/h)")
        ws.cell(row=data_start, column=aux_col + 1, value=f"{spd}% speed H (m)")
        for i, pt in enumerate(pts):
            ws.cell(row=data_start + 1 + i, column=aux_col,     value=pt.get("Q_m3h", 0))
            ws.cell(row=data_start + 1 + i, column=aux_col + 1, value=pt.get("value",  0))
        n_pts = len(pts)
        spd_series.append((aux_col, aux_col + 1, data_start, data_start + n_pts, f"{spd}%"))
        aux_col += 3

    # ── write duty-point auxiliary column for marker series ──────────
    op_col = aux_col
    ws.cell(row=data_start, column=op_col,     value="OP Q (m³/h)")
    ws.cell(row=data_start, column=op_col + 1, value="OP H (m) [Duty point]")
    for i, op in enumerate(ops[:4]):       # max 4 operating points
        ws.cell(row=data_start + 1 + i, column=op_col,     value=op.get("Q_m3h", 0))
        ws.cell(row=data_start + 1 + i, column=op_col + 1, value=op.get("H_m",   0))

    # ── Operating points table below data ────────────────────────────────
    if ops:
        row += 1
        _hdr(ws, row, 1, "N pumps",      TEAL_HDR)
        _hdr(ws, row, 2, "Q_op (m³/h)", TEAL_HDR)
        _hdr(ws, row, 3, "H_op (m)",    TEAL_HDR)
        row += 1
        for op in ops:
            _dat(ws, row, 1, op.get("n_pumps", 1), align="center")
            _dat(ws, row, 2, op.get("Q_m3h",   0), fmt=FMT_FLOW, align="right")
            _dat(ws, row, 3, op.get("H_m",     0), fmt=FMT_HEAD, align="right")
            row += 1

    # ── Chart: system curve + rated H-Q + per-speed overlays + duty marker ─
    chart = LineChart()
    chart.title  = "Pump H-Q Curves vs System Curve"
    chart.style  = 10
    chart.y_axis.title = "Head (m)"
    chart.x_axis.title = "Flow Q (m³/h)"
    chart.height = 18
    chart.width  = 28

    x_ref   = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_end)
    sys_ref = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
    hq_ref  = Reference(ws, min_col=3, min_row=data_start, max_row=data_end)
    chart.add_data(sys_ref, titles_from_data=True)
    chart.add_data(hq_ref,  titles_from_data=True)
    chart.set_categories(x_ref)

    # Style first two series (system curve = red, rated H-Q = navy)
    _series_colors = ["E74C3C", "1F3864", "27AE60", "8E44AD", "F39C12", "2980B9"]
    for idx, clr in enumerate(_series_colors[:2]):
        if idx < len(chart.series):
            chart.series[idx].graphicalProperties.line.solidFill = clr
            chart.series[idx].graphicalProperties.line.width = 22000

    # Add per-speed curve series — do NOT call set_categories again; the
    # chart keeps the categories set above (main Q column).  Calling
    # set_categories() a second time overwrites the x-axis for ALL series.
    for si, (col_q, col_h, r_s, r_e, lbl) in enumerate(spd_series):
        spd_ref = Reference(ws, min_col=col_h, min_row=r_s, max_row=r_e)
        chart.add_data(spd_ref, titles_from_data=True)
        clr_idx = 2 + si
        if clr_idx < len(chart.series):
            chart.series[clr_idx].graphicalProperties.line.solidFill = _series_colors[clr_idx % len(_series_colors)]
            chart.series[clr_idx].graphicalProperties.line.width = 14000

    # Add duty-point as a marker-only series
    if ops:
        n_op = min(len(ops), 4)
        op_ref = Reference(ws, min_col=op_col + 1,
                           min_row=data_start, max_row=data_start + n_op)
        chart.add_data(op_ref, titles_from_data=True)
        last_idx = len(chart.series) - 1
        if last_idx >= 0:
            chart.series[last_idx].graphicalProperties.line.noFill = True
            chart.series[last_idx].graphicalProperties.line.width  = 0

    ws.add_chart(chart, f"E{data_start}")
    _col_widths(ws, [14, 14, 16])


# ---------------------------------------------------------------------------
# Sheet 4 — Pump Curves  (chart: η vs Q per speed)
# ---------------------------------------------------------------------------

def _sh_pump_curves(wb: Workbook, draft: dict) -> None:
    ws   = wb.create_sheet("Pump Curves")
    meta = draft.get("meta", {}) or {}

    _title_banner(ws, "Pump Performance Curves",
                  "Q/H/η/P/NPSHr — rated (duty) speed highlighted; VFD speed sections below")
    row = 3
    row = _meta_rows(ws, meta, row)

    pr = draft.get("pumpResult") or {}
    if not pr:
        _no_data(ws, row)
        return

    hq_pts    = pr.get("hq_curve",    [])
    eta_pts   = pr.get("eta_curve",   [])
    p_pts     = pr.get("p_curve",     [])
    npshr_pts = pr.get("npshr_curve", [])
    spd_curves = pr.get("speed_curves", [])

    # ── Rated (duty) speed — highlighted in gold/NAVY ───────────────────
    _section_hdr(ws, row, "Rated Speed — 100% (Duty Speed)", 6)
    # Override section fill to distinguish duty speed
    ws.cell(row=row, column=1).fill = _fill(NAVY)
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFD700")
    row += 1
    data_start = row
    col_hdrs = ["Q (m³/h)", "H (m)", "η (%)", "P (kW)", "NPSHr (m)"]
    for ci, h in enumerate(col_hdrs, 1):
        _hdr(ws, row, ci, h, BLUE_MID)
    row += 1
    ws.freeze_panes = f"A{row}"

    n = max(len(hq_pts), len(eta_pts), len(p_pts), len(npshr_pts))
    for i in range(n):
        Q_v = hq_pts[i].get("Q_m3h",   "") if i < len(hq_pts)    else ""
        H_v = hq_pts[i].get("value",   "") if i < len(hq_pts)    else ""
        e_v = eta_pts[i].get("value",  "") if i < len(eta_pts)   else ""
        p_v = p_pts[i].get("value",    "") if i < len(p_pts)     else ""
        n_v = npshr_pts[i].get("value","") if i < len(npshr_pts) else ""
        alt = (i % 2 == 1)
        fmts = [FMT_FLOW, FMT_HEAD, FMT_EFF, FMT_POW, FMT_NPSH]
        for ci, (v, fmt) in enumerate(zip([Q_v, H_v, e_v, p_v, n_v], fmts), 1):
            _dat(ws, row, ci, v,
                 fmt=fmt if isinstance(v, float) else None,
                 align="right" if isinstance(v, (int, float)) else "left",
                 alt=alt)
        row += 1

    data_end = row - 1

    # ── Variable speed sections ──────────────────────────────────────────
    if spd_curves:
        row += 1
        _section_hdr(ws, row, "Variable Speed Curves — Affinity-law scaled", 6)
        row += 1
        for sc in spd_curves:
            spd   = sc.get("speed_pct", 100)
            ratio = spd / 100.0
            pts   = sc.get("hq_pts", [])
            # Section header per speed
            _section_hdr(ws, row, f"  {spd}% Speed", 6)
            row += 1
            for ci, h in enumerate(col_hdrs, 1):
                _hdr(ws, row, ci, h, BLUE_LITE, NAVY, 9)
            row += 1
            # Build lookup maps from rated curves (for affinity-law scaling)
            eta_map   = {pt.get("Q_m3h", 0): pt.get("value") for pt in eta_pts}
            p_map     = {pt.get("Q_m3h", 0): pt.get("value") for pt in p_pts}
            npshr_map = {pt.get("Q_m3h", 0): pt.get("value") for pt in npshr_pts}

            for i, pt in enumerate(pts):
                alt  = (i % 2 == 1)
                Q_sc = pt.get("Q_m3h", 0)
                H_sc = pt.get("value",  0)
                # Inverse-map Q to rated Q (Q_rated = Q_sc / ratio)
                Q_0  = Q_sc / ratio if ratio > 0 else Q_sc
                # η ≈ constant (affinity law); interpolate nearest rated point
                eta_v    = _nearest(eta_map,   Q_0)
                p_rated  = _nearest(p_map,     Q_0)
                np_rated = _nearest(npshr_map, Q_0)
                P_sc  = p_rated  * ratio ** 3 if p_rated  is not None else None
                N_sc  = np_rated * ratio ** 2 if np_rated is not None else None
                fmts2 = [FMT_FLOW, FMT_HEAD, FMT_EFF, FMT_POW, FMT_NPSH]
                for ci, (v, fmt) in enumerate(
                        zip([Q_sc, H_sc, eta_v, P_sc, N_sc], fmts2), 1):
                    _dat(ws, row, ci, v,
                         fmt=fmt if isinstance(v, float) else None,
                         align="right" if isinstance(v, (int, float)) else "left",
                         alt=alt)
                row += 1
            row += 1   # blank row between speeds

    # ── Chart: η vs Q (rated speed) ─────────────────────────────────────
    if eta_pts:
        chart = LineChart()
        chart.title  = "Pump Efficiency vs Flow — Rated Speed"
        chart.style  = 10
        chart.y_axis.title = "Efficiency η (%)"
        chart.x_axis.title = "Flow Q (m³/h)"
        chart.height = 14
        chart.width  = 22

        eta_ref = Reference(ws, min_col=3, min_row=data_start, max_row=data_end)
        x_ref   = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_end)
        chart.add_data(eta_ref, titles_from_data=True)
        chart.set_categories(x_ref)
        chart.series[0].graphicalProperties.line.solidFill = "27AE60"
        chart.series[0].graphicalProperties.line.width = 22000

        ws.add_chart(chart, f"G{data_start}")

    _col_widths(ws, [12, 12, 10, 10, 12])


# ---------------------------------------------------------------------------
# Sheet 5 — Operating Points  (charts: P vs Q, NPSHa vs NPSHr)
# ---------------------------------------------------------------------------

def _sh_operating_points(wb: Workbook, draft: dict) -> None:
    ws   = wb.create_sheet("Operating Points")
    meta = draft.get("meta", {}) or {}

    _title_banner(ws, "Pump Operating Points",
                  "Intersection of pump H-Q and system curves; NPSH check")
    row = 3
    row = _meta_rows(ws, meta, row)

    pr = draft.get("pumpResult") or {}
    if not pr:
        _no_data(ws, row)
        return

    ops      = pr.get("operating_points", [])
    p_pts    = pr.get("p_curve",     [])
    npshr_pts = pr.get("npshr_curve", [])

    # Operating points table
    data_start = row
    hdrs = ["Pumps", "Q (m³/h)", "Q (L/s)", "H (m)", "η (%)",
            "P (kW)", "NPSHr (m)", "NPSHa (m)", "Margin (m)", "Status"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, row, ci, h, BLUE_MID)
    row += 1

    for i, op in enumerate(ops):
        Q_v  = op.get("Q_m3h",        0)
        alt  = (i % 2 == 1)
        vals = [
            op.get("n_pumps",       1),
            Q_v,
            round(Q_v / 3.6, 3) if Q_v else "",
            op.get("H_m",           0),
            op.get("eta_pct"),
            op.get("power_kW"),
            op.get("npshr_m"),
            op.get("npsha_m"),
            op.get("npsh_margin_m"),
            "OK" if not op.get("warnings") else f"WARN: {op['warnings'][0][:40]}",
        ]
        for ci, v in enumerate(vals, 1):
            _dat(ws, row, ci, v,
                 fmt="#,##0.00" if isinstance(v, float) else None,
                 align="right" if isinstance(v, (int, float)) else "left",
                 alt=alt)
        row += 1

    data_end = row - 1

    # Write P and NPSHr/NPSHa curves for charts in helper columns
    aux_col = 13
    ws.cell(row=data_start, column=aux_col,     value="Q_p")
    ws.cell(row=data_start, column=aux_col + 1, value="P (kW)")
    ws.cell(row=data_start, column=aux_col + 2, value="Q_n")
    ws.cell(row=data_start, column=aux_col + 3, value="NPSHr (m)")
    ws.cell(row=data_start, column=aux_col + 4, value="NPSHa (m)")

    p_n    = len(p_pts)
    npshr_n = len(npshr_pts)
    for i, pt in enumerate(p_pts):
        ws.cell(row=data_start + 1 + i, column=aux_col,     value=pt.get("Q_m3h", 0))
        ws.cell(row=data_start + 1 + i, column=aux_col + 1, value=pt.get("value",  0))

    for i, pt in enumerate(npshr_pts):
        ws.cell(row=data_start + 1 + i, column=aux_col + 2, value=pt.get("Q_m3h", 0))
        ws.cell(row=data_start + 1 + i, column=aux_col + 3, value=pt.get("value",  0))
        # NPSHa from operating points (approximate by interpolation via steady value)
        # Use the first operating point NPSHa for simplicity
        npsha_v = ops[0].get("npsha_m") if ops else None
        ws.cell(row=data_start + 1 + i, column=aux_col + 4, value=npsha_v)

    chart_row = data_end + 3

    # Chart 1: P vs Q
    if p_pts:
        chart1 = LineChart()
        chart1.title  = "Power vs Flow"
        chart1.style  = 10
        chart1.y_axis.title = "Power (kW)"
        chart1.x_axis.title = "Flow (m³/h)"
        chart1.height = 12
        chart1.width  = 20

        p_end = data_start + p_n
        p_ref = Reference(ws, min_col=aux_col + 1, min_row=data_start, max_row=p_end)
        x_ref = Reference(ws, min_col=aux_col,     min_row=data_start + 1, max_row=p_end)
        chart1.add_data(p_ref, titles_from_data=True)
        chart1.set_categories(x_ref)
        chart1.series[0].graphicalProperties.line.solidFill = "E74C3C"
        chart1.series[0].graphicalProperties.line.width = 20000
        ws.add_chart(chart1, f"A{chart_row}")

    # Chart 2: NPSHa vs NPSHr
    if npshr_pts:
        chart2 = LineChart()
        chart2.title  = "NPSHa vs NPSHr"
        chart2.style  = 10
        chart2.y_axis.title = "NPSH (m)"
        chart2.x_axis.title = "Flow (m³/h)"
        chart2.height = 12
        chart2.width  = 20

        n_end = data_start + npshr_n
        r_ref = Reference(ws, min_col=aux_col + 3, min_row=data_start, max_row=n_end)
        a_ref = Reference(ws, min_col=aux_col + 4, min_row=data_start, max_row=n_end)
        x_ref = Reference(ws, min_col=aux_col + 2, min_row=data_start + 1, max_row=n_end)
        chart2.add_data(r_ref, titles_from_data=True)
        chart2.add_data(a_ref, titles_from_data=True)
        chart2.set_categories(x_ref)
        chart2.series[0].graphicalProperties.line.solidFill = "E74C3C"
        chart2.series[0].graphicalProperties.line.width = 18000
        if len(chart2.series) > 1:
            chart2.series[1].graphicalProperties.line.solidFill = "27AE60"
            chart2.series[1].graphicalProperties.line.width = 18000
        ws.add_chart(chart2, f"L{chart_row}")

    _col_widths(ws, [8, 12, 10, 10, 10, 10, 12, 12, 14, 30])


# ---------------------------------------------------------------------------
# Sheet 6 — Wet Well
# ---------------------------------------------------------------------------

def _sh_wet_well(wb: Workbook, draft: dict) -> None:
    ws   = wb.create_sheet("Wet Well")
    meta = draft.get("meta", {}) or {}
    us   = draft.get("unitSystem", "SI")

    _title_banner(ws, "Wet Well / Clear Well Sizing",
                  "Geometry, volume curve, cycle analysis, detention time")
    row = 3
    row = _meta_rows(ws, meta, row)

    cw = draft.get("clearwellResult")
    cc = draft.get("clearwellConfig") or {}
    if not cw:
        _no_data(ws, row)
        return

    # KPIs
    op_vol = cw.get("operating_volume_m3")
    det    = cw.get("detention_time_min")
    req    = cw.get("required_detention_min", 0)
    op_dv, op_du = _us(us, op_vol, "volume")
    kpis = [
        ("Operating volume",     op_vol,  op_dv, op_du),
        ("Detention time",       det,     det,   "min"),
        ("Required detention",   req,     req,   "min"),
        ("Detention OK?",        "Yes" if cw.get("detention_ok") else "No", None, "—"),
    ]
    _hdr(ws, row, 1, "KPI", BLUE_MID);  _hdr(ws, row, 2, "SI value", BLUE_MID)
    _hdr(ws, row, 3, "Display", BLUE_MID); _hdr(ws, row, 4, "Unit", BLUE_MID)
    row += 1
    for i, (lbl, si_v, dv, du) in enumerate(kpis):
        alt = (i % 2 == 1)
        _dat(ws, row, 1, lbl, bold=True, alt=alt)
        _dat(ws, row, 2, si_v, fmt="#,##0.00" if isinstance(si_v, float) else None, align="right", alt=alt)
        _dat(ws, row, 3, dv,   fmt="#,##0.00" if isinstance(dv, float) else None, align="right", alt=alt)
        _dat(ws, row, 4, du,  align="center", alt=alt)
        row += 1

    row += 1
    _section_hdr(ws, row, "Volume Curve (Level vs Volume)")
    row += 1
    hdrs = ["Level (m)", "Depth (m)", "Volume (m³)", "Volume display"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, row, ci, h, BLUE_LITE, NAVY, 9)
    row += 1
    for i, pt in enumerate(cw.get("volume_curve", [])):
        alt = (i % 2 == 1)
        v_m3   = pt.get("volume_m3", 0)
        dv, du = _us(us, v_m3, "volume")
        _dat(ws, row, 1, pt.get("level_m", 0),   fmt="#,##0.000", align="right", alt=alt)
        _dat(ws, row, 2, pt.get("depth_m", 0),   fmt="#,##0.000", align="right", alt=alt)
        _dat(ws, row, 3, v_m3,                   fmt="#,##0.00",  align="right", alt=alt)
        _dat(ws, row, 4, f"{dv:.2f} {du}",        alt=alt)
        row += 1

    row += 1
    _section_hdr(ws, row, "Cycle Results")
    row += 1
    cr_hdrs = ["Stage", "Label", "Q_pump (m³/h)", "Q_in (m³/h)",
               "t_fill (s)", "t_drain (s)", "t_cycle (s)",
               "Cycles/hr", "V_req (m³)", "Can drain?", "Cycles OK?"]
    for ci, h in enumerate(cr_hdrs, 1):
        _hdr(ws, row, ci, h, BLUE_MID, WHITE, 9)
    row += 1
    for i, cr in enumerate(cw.get("cycle_results", [])):
        alt = (i % 2 == 1)
        vals = [
            cr.get("stage"),          cr.get("label"),
            cr.get("Q_pump_m3h"),     cr.get("Q_in_m3h"),
            cr.get("t_fill_s"),       cr.get("t_drain_s"),
            cr.get("t_cycle_s"),      cr.get("cycles_per_hour"),
            cr.get("V_req_m3"),
            "Yes" if cr.get("pump_can_drain") else "No",
            "Yes" if cr.get("cycles_ok")      else "No",
        ]
        for ci, v in enumerate(vals, 1):
            _dat(ws, row, ci, v,
                 fmt="#,##0.00" if isinstance(v, float) else None,
                 align="right" if isinstance(v, (int, float)) else "left",
                 alt=alt)
        row += 1

    _col_widths(ws, [14, 14, 14, 18, 14, 14, 14, 12, 14, 12, 12])


def _hdr_font_size(ws, row: int, col: int, h: str, bg: str, fg: str = WHITE,
                   font_size: int = 10) -> None:
    _hdr(ws, row, col, h, bg, fg, font_size)


# ---------------------------------------------------------------------------
# Sheet 7 — Engineering Checks
# ---------------------------------------------------------------------------

def _sh_checks(wb: Workbook, draft: dict) -> None:
    ws   = wb.create_sheet("Engineering Checks")
    meta = draft.get("meta", {}) or {}

    _title_banner(ws, "Engineering Checks — Compliance Summary",
                  "Velocity, Re, NPSH, cycling, surge per AS 2200 / AWWA M11 / AS 2941")
    row = 3
    row = _meta_rows(ws, meta, row)

    r   = draft.get("hydraulicsResult", {}) or {}
    pr  = draft.get("pumpResult",       {}) or {}
    cw  = draft.get("clearwellResult",  {}) or {}
    whr = draft.get("waterHammerResult",{}) or {}

    checks: list[tuple[str, str, str, str, str, str]] = []

    v = r.get("velocity_ms")
    if v is not None:
        if v < 0.6:    sev, note = "WARNING",  f"v = {v:.3f} m/s < 0.6 m/s — sedimentation risk"
        elif v > 3.0:  sev, note = "CRITICAL", f"v = {v:.3f} m/s > 3.0 m/s — excessive friction & noise"
        else:          sev, note = "OK",        f"v = {v:.3f} m/s within 0.6–3.0 m/s"
        checks.append(("Pipe velocity", f"{v:.3f} m/s", "0.6–3.0 m/s", "AS 2200", sev, note))

    Re = r.get("reynolds_number")
    if Re is not None:
        if Re > 4000:   sev, note = "OK",       f"Re = {Re:,.0f} — turbulent"
        elif Re > 2300: sev, note = "WARNING",  f"Re = {Re:,.0f} — transitional"
        else:           sev, note = "CRITICAL", f"Re = {Re:,.0f} — laminar at design Q"
        checks.append(("Reynolds number", f"{Re:,.0f}", "> 4 000", "D-W", sev, note))

    tdh = r.get("tdh_m")
    if tdh is not None:
        sev  = "WARNING" if tdh > 200 else "OK"
        note = f"TDH = {tdh:.2f} m. {'High — verify pump selection.' if tdh > 200 else 'Within normal range.'}"
        checks.append(("TDH", f"{tdh:.2f} m", "< 200 m typical", "AWWA M11", sev, note))

    for op in (pr.get("operating_points") or []):
        m = op.get("npsh_margin_m")
        if m is not None:
            sev  = "OK" if m >= 0.5 else ("WARNING" if m >= 0 else "CRITICAL")
            note = f"NPSHa margin = {m:.2f} m at {op.get('n_pumps',1)} pump(s). Min ≥ 0.5 m."
            checks.append((f"NPSH margin (N={op.get('n_pumps',1)})", f"{m:.2f} m",
                           "≥ 0.5 m", "ISO 9906", sev, note))

    for cr in (cw.get("cycle_results") or []):
        cph = cr.get("cycles_per_hour", 0)
        ok  = cr.get("cycles_ok", True)
        sev  = "OK" if ok else "WARNING"
        note = f"Stage {cr.get('stage',1)}: {cph:.1f} cycles/hr"
        checks.append((f"Cycling (stage {cr.get('stage',1)})", f"{cph:.2f}/hr",
                       "≤ 6/hr", "AWWA M37", sev, note))

    if whr.get("cavitation_risk"):
        checks.append(("Surge — cavitation risk", "YES", "NO", "AS 2941", "CRITICAL",
                        f"Min head {whr.get('min_pressure_head_m',0):.2f} m < h_vap. Column separation."))
    elif whr.get("delta_H_m") is not None:
        dH  = whr["delta_H_m"]
        sev  = "WARNING" if dH > 50 else "OK"
        note = f"ΔH = {dH:.1f} m. {'Consider surge protection.' if dH > 50 else 'Acceptable.'}"
        checks.append(("Surge — Joukowsky ΔH", f"{dH:.1f} m", "< 50 m guide", "AS 2941", sev, note))

    if not checks:
        _no_data(ws, row)
        return

    hdrs = ["Check", "Value", "Criterion", "Standard", "Result", "Notes"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, row, ci, h, BLUE_MID)
    ws.merge_cells(f"F{row}:L{row}")
    row += 1

    for i, (check, val, crit, std, sev, note) in enumerate(checks):
        alt = (i % 2 == 1)
        _dat(ws, row, 1, check, bold=True, alt=alt)
        _dat(ws, row, 2, val,   align="right", alt=alt)
        _dat(ws, row, 3, crit,  alt=alt)
        _dat(ws, row, 4, std,   alt=alt)
        c5 = ws.cell(row=row, column=5, value=sev)
        c5.fill      = _fill(SEV_COLORS.get(sev, BLUE_MID))
        c5.font      = Font(bold=True, size=10, color=WHITE)
        c5.alignment = Alignment(horizontal="center")
        c5.border    = _border()
        ws.merge_cells(f"F{row}:L{row}")
        _dat(ws, row, 6, note, alt=alt)
        row += 1

    _col_widths(ws, [30, 16, 22, 12, 12, 60])


# ---------------------------------------------------------------------------
# Sheet 8 — Surge Quick (suction + discharge sections)
# ---------------------------------------------------------------------------

def _surge_quick_section(ws, row: int, label: str, whr: dict, us: str) -> int:
    """Write one pipeline's surge quick results. Return next row."""
    _section_hdr(ws, row, label, 5)
    row += 1
    if not whr:
        _no_data(ws, row, "No result for this pipeline.")
        return row + 2

    params = [
        ("Pipeline",                    whr.get("pipeline",           ""),  None),
        ("Event type",                  whr.get("event_type",         ""),  None),
        ("Wave speed a",                whr.get("wave_speed_ms",       0),  "m/s"),
        ("Steady velocity V₀",          whr.get("V0_ms",               0),  "m/s"),
        ("Pipe length L",               whr.get("pipe_length_m",       0),  "m"),
        ("Fluid density ρ",             whr.get("rho_kg_m3",        1000),  "kg/m³"),
        ("Operating head H₀",           whr.get("H_operating_m",       0),  "m"),
        ("Joukowsky ΔH",                whr.get("delta_H_joukowsky_m", 0),  "m"),
        ("Joukowsky ΔP",                whr.get("delta_P_joukowsky_kPa",0), "kPa"),
        ("Pipe period T = 2L/a",        whr.get("T_char_s",            0),  "s"),
        ("Closure time t_c",            whr.get("closure_time_s"),         "s"),
        ("Reduction factor K",          whr.get("reduction_factor",    1.0),"—"),
        ("Effective ΔH",                whr.get("delta_H_m",           0),  "m"),
        ("Effective ΔP",                whr.get("delta_P_kPa",         0),  "kPa"),
        ("Max pressure head",           whr.get("max_pressure_head_m", 0),  "m"),
        ("Min pressure head",           whr.get("min_pressure_head_m", 0),  "m"),
        ("Cavitation risk",             "YES" if whr.get("cavitation_risk") else "NO", None),
        ("Vapor pressure head h_vap",   whr.get("vapor_pressure_head_m",0), "m"),
        ("Temperature",                 whr.get("temperature_C",       20), "°C"),
    ]
    rc = whr.get("rating_check") or {}
    if rc:
        params += [
            ("Rating status",          rc.get("rating_status","").upper(), None),
            ("Pressure class (PN)",    rc.get("pressure_rating_kPa", 0),   "kPa"),
            ("Max transient P",        rc.get("max_transient_kPa",   0),   "kPa"),
            ("FoS = PN / P_max",       rc.get("factor_of_safety",    0),   "—"),
        ]

    _hdr(ws, row, 1, "Parameter", BLUE_LITE, NAVY)
    _hdr(ws, row, 2, "Value",     BLUE_LITE, NAVY)
    _hdr(ws, row, 3, "Unit",      BLUE_LITE, NAVY)
    row += 1
    for i, (p, v, u) in enumerate(params):
        alt = (i % 2 == 1)
        _dat(ws, row, 1, p, bold=True, alt=alt)
        _dat(ws, row, 2, v, fmt="#,##0.0000" if isinstance(v, float) else None,
             align="right", alt=alt)
        _dat(ws, row, 3, u or "—", align="center", alt=alt)
        row += 1

    return row + 1


def _sh_surge_quick(wb: Workbook, draft: dict) -> None:
    ws  = wb.create_sheet("Surge Quick (Mode A)")
    meta = draft.get("meta", {}) or {}
    us   = draft.get("unitSystem", "SI")

    _title_banner(ws, "Water Hammer — Joukowsky / Allievi Quick Analysis",
                  "Mode A: suction and discharge pipelines")
    row = 3
    row = _meta_rows(ws, meta, row)

    whr = draft.get("waterHammerResult") or {}
    pipeline = whr.get("pipeline", "discharge") if whr else "discharge"

    suction_whr  = whr if (pipeline == "suction") else {}
    discharge_whr = whr if (pipeline == "discharge") else {}

    row = _surge_quick_section(ws, row, "SUCTION PIPELINE — Quick Surge", suction_whr, us)
    row = _surge_quick_section(ws, row, "DISCHARGE PIPELINE — Quick Surge", discharge_whr, us)

    _col_widths(ws, [34, 18, 10])


# ---------------------------------------------------------------------------
# Sheet 9 — Surge MOC Time Histories
# ---------------------------------------------------------------------------

MAX_HIST_ROWS = 1000


def _write_pipeline_history(ws, row: int, label: str,
                             moc: dict | None) -> tuple[int, list]:
    """Write observation point histories for one pipeline. Returns (next_row, chart_refs)."""
    _section_hdr(ws, row, label, 12)
    row += 1
    if not moc:
        _no_data(ws, row, "MOC not run for this pipeline.")
        return row + 2, []

    obs_pts = moc.get("observations", [])
    cavit_x = moc.get("cavitation_x_m", [])
    if cavit_x:
        ws.merge_cells(f"A{row}:L{row}")
        c = ws.cell(row=row, column=1, value=f"⚠ Column separation at: {cavit_x[:5]}")
        c.font = Font(bold=True, color="E74C3C", size=10)
        row += 1

    if not obs_pts:
        _no_data(ws, row, "No observation points in MOC result.")
        return row + 2, []

    chart_refs: list[tuple[int, int, int, str]] = []
    col_base = 1
    header_row = row

    for obs in obs_pts:
        lbl     = obs.get("label", "Obs")
        history = _subsample(obs.get("history", []), MAX_HIST_ROWS)
        n_rows  = len(history)
        if not history:
            continue

        ws.cell(row=header_row, column=col_base,     value=f"{lbl} — t (s)").font = Font(bold=True, size=9, color=WHITE)
        ws.cell(row=header_row, column=col_base).fill = _fill(BLUE_MID)
        ws.cell(row=header_row, column=col_base).border = _border()
        ws.cell(row=header_row, column=col_base + 1, value=f"{lbl} — H (m)").font = Font(bold=True, size=9, color=WHITE)
        ws.cell(row=header_row, column=col_base + 1).fill = _fill(BLUE_MID)
        ws.cell(row=header_row, column=col_base + 1).border = _border()

        for i, pt in enumerate(history):
            ws.cell(row=header_row + 1 + i, column=col_base,     value=pt.get("t_s", 0))
            ws.cell(row=header_row + 1 + i, column=col_base + 1, value=pt.get("H_m", 0))

        chart_refs.append((col_base, col_base + 1, header_row, header_row + n_rows, lbl))
        col_base += 3

    next_row = header_row + max(
        (len(_subsample(obs.get("history", []), MAX_HIST_ROWS)) for obs in obs_pts), default=0
    ) + 2
    return next_row, chart_refs


def _sh_moc_histories(wb: Workbook, draft: dict) -> None:
    ws  = wb.create_sheet("Surge MOC Time Histories")
    meta = draft.get("meta", {}) or {}

    _title_banner(ws, "Surge MOC — Head Time History at Observation Points",
                  "Suction pipeline (NPSHa transient) then discharge pipeline")
    row = 3
    row = _meta_rows(ws, meta, row)

    suct_moc  = draft.get("suctionSurgeResult")
    disc_moc  = draft.get("mocResult")

    row, suct_refs = _write_pipeline_history(ws, row, "SUCTION PIPELINE — MOC Time History", suct_moc)
    row, disc_refs = _write_pipeline_history(ws, row, "DISCHARGE PIPELINE — MOC Time History", disc_moc)

    all_refs = suct_refs + disc_refs
    if not all_refs:
        return

    # Single combined chart
    chart = LineChart()
    chart.title  = "Head vs Time — All Observation Points"
    chart.style  = 10
    chart.y_axis.title = "Head (m)"
    chart.x_axis.title = "Time (s)"
    chart.height = 18
    chart.width  = 28

    chart_colors = ["1F3864", "E74C3C", "27AE60", "8E44AD", "F39C12"]
    for idx, (col_t, col_H, r_start, r_end, lbl) in enumerate(all_refs[:5]):
        y_ref = Reference(ws, min_col=col_H, min_row=r_start, max_row=r_end)
        x_ref = Reference(ws, min_col=col_t, min_row=r_start + 1, max_row=r_end)
        chart.add_data(y_ref, titles_from_data=True)
        chart.set_categories(x_ref)
        if idx < len(chart.series):
            chart.series[idx].graphicalProperties.line.solidFill = chart_colors[idx % len(chart_colors)]
            chart.series[idx].graphicalProperties.line.width = 18000

    ws.add_chart(chart, f"A{row + 1}")
    _col_widths(ws, [12, 12, 2] * 6)


# ---------------------------------------------------------------------------
# Sheet 10 — Surge Envelope vs Distance
# ---------------------------------------------------------------------------

def _write_pipeline_envelope(ws, row: int, label: str,
                              moc: dict | None, us: str) -> tuple[int, int, int]:
    """Write one pipeline's envelope data. Returns (next_row, data_start, data_end)."""
    _section_hdr(ws, row, label, 8)
    row += 1
    if not moc:
        _no_data(ws, row, "MOC not run for this pipeline.")
        return row + 2, row, row

    env = moc.get("envelope", [])

    # Global KPIs
    kpis = [
        ("Global max H",  moc.get("global_max_H_m",   0), "m"),
        ("Global min H",  moc.get("global_min_H_m",   0), "m"),
        ("Cav. nodes",    len(moc.get("cavitation_x_m", [])), "—"),
    ]
    for lbl, val, unit in kpis:
        _dat(ws, row, 1, lbl, bold=True)
        _dat(ws, row, 2, val, fmt="#,##0.00" if isinstance(val, float) else None, align="right")
        _dat(ws, row, 3, unit, align="center")
        row += 1
    row += 1

    if not env:
        _no_data(ws, row, "No envelope data.")
        return row + 2, row, row

    data_start = row
    hdrs = ["x (m)", "Elev (m)", "H_max (m)", "H_min (m)", "P_max (kPa)", "P_min (kPa)"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, row, ci, h, BLUE_LITE, NAVY, 9)
    row += 1

    for i, pt in enumerate(env):
        alt = (i % 2 == 1)
        vals = [pt.get("x_m",0), pt.get("elev_m",0),
                pt.get("H_max_m",0), pt.get("H_min_m",0),
                pt.get("P_max_kPa",0), pt.get("P_min_kPa",0)]
        for ci, v in enumerate(vals, 1):
            _dat(ws, row, ci, v, fmt="#,##0.00", align="right", alt=alt)
        row += 1

    return row + 1, data_start, row - 1


def _sh_surge_envelope(wb: Workbook, draft: dict) -> None:
    ws  = wb.create_sheet("Surge Envelope vs Distance")
    meta = draft.get("meta", {}) or {}
    us   = draft.get("unitSystem", "SI")

    _title_banner(ws, "Surge Pressure Envelope — H_max / H_min vs Distance",
                  "Suction pipeline then discharge pipeline (Method of Characteristics)")
    row = 3
    row = _meta_rows(ws, meta, row)

    suct_moc = draft.get("suctionSurgeResult")
    disc_moc = draft.get("mocResult")

    row, suct_start, suct_end = _write_pipeline_envelope(ws, row, "SUCTION PIPELINE", suct_moc, us)
    row, disc_start, disc_end = _write_pipeline_envelope(ws, row, "DISCHARGE PIPELINE", disc_moc, us)

    # Chart — suction envelope
    charts_added = 0
    for ds, de, lbl in [(suct_start, suct_end, "Suction"),
                        (disc_start, disc_end, "Discharge")]:
        if de <= ds:
            continue
        chart = LineChart()
        chart.title  = f"Surge Pressure Envelope — {lbl}"
        chart.style  = 10
        chart.y_axis.title = "Head (m)"
        chart.x_axis.title = "Distance (m)"
        chart.height = 14
        chart.width  = 22

        hmax_ref = Reference(ws, min_col=3, min_row=ds, max_row=de)
        hmin_ref = Reference(ws, min_col=4, min_row=ds, max_row=de)
        x_ref    = Reference(ws, min_col=1, min_row=ds + 1, max_row=de)
        chart.add_data(hmax_ref, titles_from_data=True)
        chart.add_data(hmin_ref, titles_from_data=True)
        chart.set_categories(x_ref)
        chart.series[0].graphicalProperties.line.solidFill = "E74C3C"
        chart.series[0].graphicalProperties.line.width = 20000
        if len(chart.series) > 1:
            chart.series[1].graphicalProperties.line.solidFill = "2980B9"
            chart.series[1].graphicalProperties.line.width = 20000

        # Suction chart anchored at col H; discharge chart anchored at col P
        anchor_col = "H" if charts_added == 0 else "P"
        anchor_row = ds
        ws.add_chart(chart, f"{anchor_col}{anchor_row}")
        charts_added += 1
        # Do NOT break — plot both suction AND discharge pipelines

    _col_widths(ws, [12, 12, 14, 14, 16, 16])


# ---------------------------------------------------------------------------
# Sheet 11 — Protection Comparisons
# ---------------------------------------------------------------------------

def _sh_protection(wb: Workbook, draft: dict) -> None:
    ws  = wb.create_sheet("Protection Comparisons")
    meta = draft.get("meta", {}) or {}

    _title_banner(ws, "Surge Protection — What-If Device Comparison",
                  "Baseline + protection devices (air vessel, surge tank, PRV, vacuum relief, slow check valve)")
    row = 3
    row = _meta_rows(ws, meta, row)

    wi = draft.get("whatIfResult")
    if not wi:
        _no_data(ws, row)
        return

    baseline = wi.get("baseline", {})
    runs     = wi.get("device_runs", [])
    all_runs = [baseline] + runs

    hdrs = ["Scenario", "Max H (m)", "Min H (m)", "Max P (kPa)", "Min P (kPa)",
            "Surge ΔH (m)", "Surge Δ%", "Cav. nodes", "Cav. risk", "FoS", "Rating"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, row, ci, h, BLUE_MID)
    row += 1

    for i, run in enumerate(all_runs):
        alt = (i % 2 == 1)
        rc  = run.get("rating_check") or {}
        if run.get("run_error"):
            _dat(ws, row, 1, run.get("label", f"Run {i}"), bold=True, alt=alt)
            _dat(ws, row, 2, f"ERROR — {run['run_error']}", alt=alt, merge_to=11)
            row += 1
            continue

        vals = [
            run.get("label",                  "Baseline"),
            run.get("global_max_H_m",         0),
            run.get("global_min_H_m",         0),
            run.get("global_max_P_kPa",       0),
            run.get("global_min_P_kPa",       0),
            run.get("max_surge_reduction_m"),
            run.get("max_surge_reduction_pct"),
            len(run.get("cavitation_x_m", [])),
            "YES" if run.get("cavitation_risk") else "NO",
            rc.get("factor_of_safety") if rc else None,
            rc.get("rating_status", "N/A").upper() if rc else "N/A",
        ]
        for ci, v in enumerate(vals, 1):
            _dat(ws, row, ci, v,
                 fmt="#,##0.00" if isinstance(v, float) else None,
                 align="right" if isinstance(v, (int, float)) else "left",
                 alt=alt)
        row += 1

    notes = wi.get("assumption_notes", [])
    if notes:
        row += 1
        _section_hdr(ws, row, "Assumption Notes", 11)
        row += 1
        for note in notes:
            ws.merge_cells(f"A{row}:K{row}")
            _dat(ws, row, 1, f"• {note}")
            row += 1

    _col_widths(ws, [28, 14, 14, 14, 14, 16, 12, 12, 12, 10, 12])


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def _wb_to_bytes(wb: Workbook) -> bytes:
    """Serialise an openpyxl Workbook to raw .xlsx bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_workbook(draft: dict) -> Workbook:
    """
    Build a professional .xlsx workbook from a serialised ProjectDraft dict.

    Returns an openpyxl.Workbook object.  Callers that need raw bytes should
    call ``_wb_to_bytes(build_workbook(draft))``.
    """
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    if default is not None:
        wb.remove(default)

    # Set workbook title metadata from project name
    meta = draft.get("meta") or {}
    project_name = meta.get("name", "WPS Designer Project") or "WPS Designer Project"
    wb.properties.title   = project_name
    wb.properties.subject = "ALLL WPS Designer — Hydraulic Calculation Report"
    wb.properties.creator = meta.get("engineer", "WPS Designer")
    wb.properties.company = meta.get("client", "")

    _sh_inputs(wb, draft)
    _sh_hydraulics(wb, draft)
    _sh_system_curve(wb, draft)
    _sh_pump_curves(wb, draft)
    _sh_operating_points(wb, draft)
    _sh_wet_well(wb, draft)
    _sh_checks(wb, draft)
    _sh_surge_quick(wb, draft)
    _sh_moc_histories(wb, draft)
    _sh_surge_envelope(wb, draft)
    _sh_protection(wb, draft)

    return wb
