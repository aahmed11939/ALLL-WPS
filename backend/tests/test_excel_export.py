"""
Tests for backend/engine/excel_export.py and the POST /export/excel endpoint.

Covers:
  - build_workbook() with empty / minimal / full drafts
  - All 11 sheets are present
  - Returned bytes are a valid .xlsx file (openpyxl can open them)
  - Endpoint returns 200 with correct content-type
  - Endpoint returns 400 for invalid JSON
  - Workbook with partial results (only hydraulics, no pump/clearwell/surge)
  - Project name sanitisation in Content-Disposition header
  - Each sheet builder does not raise on missing data
  - Charts are present on the expected sheets
"""

from __future__ import annotations

import io
import json

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from openpyxl import Workbook as OpenpyxlWorkbook

from backend.api.main import app
from backend.engine.excel_export import _wb_to_bytes, build_workbook

client = TestClient(app)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

EMPTY_DRAFT: dict = {
    "meta": {
        "name": "Test Project",
        "client": "ACME Water",
        "job_number": "WPS-001",
        "date": "2026-05-07",
        "engineer": "J. Smith",
        "notes": "",
    },
    "unitSystem": "SI",
    "designFlow_m3h": 36.0,
    "upstreamNode":   {"elevation_m": 5.0,  "pressure_kPa": 0},
    "downstreamNode": {"elevation_m": 35.0, "pressure_kPa": 0},
    "suction": {
        "segments": [{"material": "pvc", "diameter_mm": 150, "length_m": 100}],
        "accessories": [],
        "accessories_K_sum": 0.5,
    },
    "discharge": {
        "segments": [{"material": "dicl", "diameter_mm": 200, "length_m": 400}],
        "accessories": [],
        "accessories_K_sum": 1.2,
    },
    "hydraulicsResult": None,
    "pumpResult":       None,
    "clearwellResult":  None,
    "waterHammerResult": None,
    "mocResult":        None,
    "suctionSurgeResult": None,
    "whatIfResult":     None,
}


HYDRAULICS_RESULT = {
    "design_Q_m3h":    36.0,
    "velocity_ms":     1.35,
    "reynolds_number": 225000.0,
    "friction_factor": 0.0165,
    "K_sum":           1.5,
    "static_head_m":   30.0,
    "friction_head_m": 8.4,
    "minor_head_m":    0.14,
    "tdh_m":           38.54,
    "system_curve": [
        {"Q_m3h":  0.0, "H_m": 30.0},
        {"Q_m3h": 18.0, "H_m": 32.0},
        {"Q_m3h": 36.0, "H_m": 38.5},
        {"Q_m3h": 54.0, "H_m": 49.5},
        {"Q_m3h": 72.0, "H_m": 65.0},
    ],
}

PUMP_RESULT = {
    "active": True,
    "hq_curve": [
        {"Q_m3h": 0.0,  "value": 42.0},
        {"Q_m3h": 36.0, "value": 38.5},
        {"Q_m3h": 72.0, "value": 28.0},
    ],
    "eta_curve": [
        {"Q_m3h": 0.0,  "value": 0.0},
        {"Q_m3h": 36.0, "value": 72.0},
        {"Q_m3h": 72.0, "value": 55.0},
    ],
    "p_curve": [
        {"Q_m3h": 0.0,  "value": 0.0},
        {"Q_m3h": 36.0, "value": 5.3},
        {"Q_m3h": 72.0, "value": 9.8},
    ],
    "npshr_curve": [
        {"Q_m3h": 0.0,  "value": 2.0},
        {"Q_m3h": 36.0, "value": 3.5},
        {"Q_m3h": 72.0, "value": 5.2},
    ],
    "speed_curves": [],
    "operating_points": [
        {
            "n_pumps":        1,
            "Q_m3h":          36.2,
            "H_m":            38.4,
            "eta_pct":        71.8,
            "power_kW":       5.3,
            "npshr_m":        3.5,
            "npsha_m":        4.5,
            "npsh_margin_m":  1.0,
            "warnings":       [],
        }
    ],
    "non_physical_fit": False,
    "warnings": [],
}

CLEARWELL_RESULT = {
    "active": True,
    "volume_curve": [
        {"level_m": 0.0, "depth_m": 0.0, "volume_m3": 0.0},
        {"level_m": 1.0, "depth_m": 1.0, "volume_m3": 19.6},
        {"level_m": 2.0, "depth_m": 2.0, "volume_m3": 39.3},
        {"level_m": 3.0, "depth_m": 3.0, "volume_m3": 58.9},
    ],
    "operating_volume_m3": 39.3,
    "cycle_results": [
        {
            "stage": 1,
            "label": "Duty",
            "Q_pump_m3h": 72.0,
            "Q_in_m3h": 36.0,
            "t_fill_s":  None,
            "t_drain_s": 1965.6,
            "t_cycle_s": 2944.0,
            "cycles_per_hour": 1.22,
            "V_req_m3": 19.6,
            "cycles_ok": True,
            "pump_can_drain": True,
        }
    ],
    "detention_time_min": 65.5,
    "required_detention_min": 30.0,
    "detention_ok": True,
    "warnings": [],
}

SURGE_QUICK_RESULT = {
    "pipeline":              "discharge",
    "event_type":            "pump_trip",
    "wave_speed_ms":         1000.0,
    "V0_ms":                 1.35,
    "pipe_length_m":         400.0,
    "rho_kg_m3":             1000.0,
    "H_operating_m":         38.5,
    "delta_V_ms":            1.35,
    "delta_H_joukowsky_m":   137.7,
    "delta_P_joukowsky_kPa": 1377.0,
    "T_char_s":              0.8,
    "closure_time_s":        None,
    "reduction_factor":      1.0,
    "reduction_method":      "instantaneous",
    "delta_H_m":             137.7,
    "delta_P_kPa":           1377.0,
    "envelope":              [],
    "min_pressure_head_m":  -99.2,
    "max_pressure_head_m":  176.2,
    "min_pressure_kPa":     -973.0,
    "max_pressure_kPa":     1730.0,
    "cavitation_risk":       True,
    "vacuum_risk":           True,
    "vapor_pressure_head_m": -10.3,
    "temperature_C":         20.0,
    "rating_check":          None,
    "unit_system":           "SI",
}

MOC_RESULT = {
    "pipeline":          "discharge",
    "N":                 10,
    "dx_m":              40.0,
    "dt_s":              0.04,
    "courant":           1.0,
    "t_total_s":         2.0,
    "n_steps":           50,
    "D_m":               0.2,
    "f":                 0.018,
    "T_char_s":          0.8,
    "envelope": [
        {"x_m": 0.0,  "elev_m": 35.0, "H_max_m": 50.0, "H_min_m": 10.0, "P_max_kPa": 147.0, "P_min_kPa": -245.0},
        {"x_m": 200.0,"elev_m": 20.0, "H_max_m": 45.0, "H_min_m":  5.0, "P_max_kPa": 245.0, "P_min_kPa": -147.0},
        {"x_m": 400.0,"elev_m":  5.0, "H_max_m": 40.0, "H_min_m":  0.0, "P_max_kPa": 343.0, "P_min_kPa":    0.0},
    ],
    "observations": [
        {
            "label":      "Pump outlet",
            "frac":       0.0,
            "node_index": 0,
            "x_m":        0.0,
            "history": [
                {"t_s": 0.0,  "H_m": 38.5, "P_kPa": 328.0},
                {"t_s": 0.04, "H_m": 42.1, "P_kPa": 373.0},
                {"t_s": 0.08, "H_m": 38.0, "P_kPa": 323.0},
            ],
        }
    ],
    "global_max_H_m":    50.0,
    "global_min_H_m":    0.0,
    "global_max_P_kPa":  343.0,
    "global_min_P_kPa":  0.0,
    "cavitation_x_m":    [400.0],
    "h_vap_m":           -10.3,
    "temperature_C":     20.0,
    "assumption_notes":  [],
    "rating_check":      None,
    "unit_system":       "SI",
}

WHATIF_RESULT = {
    "baseline": {
        "label":                   "Baseline (no device)",
        "run_error":               None,
        "global_max_H_m":          50.0,
        "global_min_H_m":          0.0,
        "global_max_P_kPa":        343.0,
        "global_min_P_kPa":        0.0,
        "max_surge_reduction_m":   None,
        "max_surge_reduction_pct": None,
        "min_head_improvement_m":  None,
        "cavitation_x_m":          [400.0],
        "cavitation_risk":         True,
        "risk_duration_s":         0.2,
        "envelope_reduction_pct":  None,
        "rating_check":            None,
        "sizing_summary":          None,
        "envelope":                [],
    },
    "device_runs": [
        {
            "label":                   "Air Vessel 0.5 m³",
            "run_error":               None,
            "global_max_H_m":          42.0,
            "global_min_H_m":          5.0,
            "global_max_P_kPa":        285.0,
            "global_min_P_kPa":        49.0,
            "max_surge_reduction_m":   8.0,
            "max_surge_reduction_pct": 16.0,
            "min_head_improvement_m":  5.0,
            "cavitation_x_m":          [],
            "cavitation_risk":         False,
            "risk_duration_s":         0.0,
            "envelope_reduction_pct":  16.0,
            "rating_check":            None,
            "sizing_summary":          None,
            "envelope":                [],
        }
    ],
    "assumption_notes": ["Screening-level analysis. ±30–50 % accuracy."],
    "t_total_s":        2.0,
    "T_char_s":         0.8,
    "pipeline":         "discharge",
}


def full_draft() -> dict:
    d = dict(EMPTY_DRAFT)
    d["hydraulicsResult"]  = HYDRAULICS_RESULT
    d["pumpResult"]        = PUMP_RESULT
    d["clearwellResult"]   = CLEARWELL_RESULT
    d["waterHammerResult"] = SURGE_QUICK_RESULT
    d["mocResult"]         = MOC_RESULT
    d["whatIfResult"]      = WHATIF_RESULT
    return d


# ===========================================================================
# Unit tests — build_workbook()
# ===========================================================================

EXPECTED_SHEETS = [
    "Inputs Summary",
    "Hydraulics Breakdown",
    "System Curve",
    "Pump Curves",
    "Operating Points",
    "Wet Well",
    "Engineering Checks",
    "Surge Quick (Mode A)",
    "Surge MOC Time Histories",
    "Surge Envelope vs Distance",
    "Protection Comparisons",
]


def _bwb(draft: dict) -> bytes:
    """Test helper: build workbook and serialize to bytes."""
    return _wb_to_bytes(build_workbook(draft))


def _cells(ws) -> list[str]:
    """Flatten all cell values in a worksheet to a list of strings."""
    return [
        str(ws.cell(r, c).value if ws.cell(r, c).value is not None else "")
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]


class TestBuildWorkbook:
    def test_returns_workbook(self):
        """build_workbook() now returns an openpyxl.Workbook, not bytes."""
        wb = build_workbook(EMPTY_DRAFT)
        assert isinstance(wb, OpenpyxlWorkbook)
        data = _wb_to_bytes(wb)
        assert len(data) > 1000  # real xlsx, not empty

    def test_is_valid_xlsx(self):
        data = _bwb(EMPTY_DRAFT)
        wb   = load_workbook(io.BytesIO(data))
        assert wb is not None

    def test_all_11_sheets_present_empty_draft(self):
        data = _bwb(EMPTY_DRAFT)
        wb   = load_workbook(io.BytesIO(data))
        assert set(EXPECTED_SHEETS).issubset(set(wb.sheetnames))

    def test_all_11_sheets_present_full_draft(self):
        data = _bwb(full_draft())
        wb   = load_workbook(io.BytesIO(data))
        assert set(EXPECTED_SHEETS).issubset(set(wb.sheetnames))

    def test_inputs_sheet_contains_project_name(self):
        d    = dict(EMPTY_DRAFT)
        d["meta"] = dict(EMPTY_DRAFT["meta"], name="My Pump Station")
        data = _bwb(d)
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Inputs Summary"]
        assert any("My Pump Station" in v for v in _cells(ws))

    def test_hydraulics_sheet_has_per_segment_data(self):
        """Hydraulics sheet always shows per-segment geometry breakdown."""
        data = _bwb(EMPTY_DRAFT)
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Hydraulics Breakdown"]
        assert any("pvc" in v.lower() or "dicl" in v.lower() for v in _cells(ws))

    def test_hydraulics_sheet_has_tdh_when_result_present(self):
        d    = dict(EMPTY_DRAFT)
        d["hydraulicsResult"] = HYDRAULICS_RESULT
        data = _bwb(d)
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Hydraulics Breakdown"]
        assert any("TDH" in v or "38.54" in v for v in _cells(ws))

    def test_system_curve_sheet_has_data_rows(self):
        d    = dict(EMPTY_DRAFT)
        d["hydraulicsResult"] = HYDRAULICS_RESULT
        data = _bwb(d)
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["System Curve"]
        q_values = [
            ws.cell(r, 1).value
            for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, (int, float))
        ]
        assert len(q_values) >= 5

    def test_pump_curves_sheet_has_hq_data(self):
        data = _bwb(full_draft())
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Pump Curves"]
        assert any("H (m)" in v or "38.5" in v for v in _cells(ws))

    def test_operating_points_sheet_has_operating_point(self):
        data = _bwb(full_draft())
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Operating Points"]
        assert any("36.2" in v or "71.8" in v for v in _cells(ws))

    def test_wet_well_sheet_has_volume_curve(self):
        data = _bwb(full_draft())
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Wet Well"]
        assert any("19.6" in v or "39.3" in v for v in _cells(ws))

    def test_engineering_checks_present(self):
        data = _bwb(full_draft())
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Engineering Checks"]
        assert any("velocity" in v.lower() or "reynolds" in v.lower() for v in _cells(ws))

    def test_surge_quick_sheet_has_wave_speed(self):
        data = _bwb(full_draft())
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Surge Quick (Mode A)"]
        assert any("1000" in v for v in _cells(ws))

    def test_moc_histories_sheet_observation_labels(self):
        data = _bwb(full_draft())
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Surge MOC Time Histories"]
        assert any("Pump outlet" in v for v in _cells(ws))

    def test_surge_envelope_sheet_has_distance_data(self):
        data = _bwb(full_draft())
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Surge Envelope vs Distance"]
        num_values = [
            ws.cell(r, 1).value
            for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, (int, float))
        ]
        assert len(num_values) >= 3

    def test_protection_comparisons_sheet_has_baseline(self):
        data = _bwb(full_draft())
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Protection Comparisons"]
        assert any("Baseline" in v or "Air Vessel" in v for v in _cells(ws))

    def test_empty_draft_no_raise(self):
        """build_workbook must not raise even when all result fields are None."""
        data = _bwb({})
        wb   = load_workbook(io.BytesIO(data))
        assert len(wb.sheetnames) == 11

    def test_system_curve_chart_attached(self):
        d    = dict(EMPTY_DRAFT)
        d["hydraulicsResult"] = HYDRAULICS_RESULT
        data = _bwb(d)
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["System Curve"]
        assert len(ws._charts) >= 1

    def test_surge_envelope_chart_attached(self):
        data = _bwb(full_draft())
        wb   = load_workbook(io.BytesIO(data))
        ws   = wb["Surge Envelope vs Distance"]
        assert len(ws._charts) >= 1


# ===========================================================================
# Integration tests — POST /export/excel endpoint
# ===========================================================================

class TestExportExcelEndpoint:
    def test_endpoint_200_with_full_draft(self):
        resp = client.post("/export/excel", json=full_draft())
        assert resp.status_code == 200

    def test_endpoint_content_type_xlsx(self):
        resp = client.post("/export/excel", json=full_draft())
        assert "spreadsheetml" in resp.headers["content-type"]

    def test_endpoint_content_disposition_has_filename(self):
        resp = client.post("/export/excel", json=full_draft())
        cd   = resp.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert ".xlsx" in cd

    def test_endpoint_filename_uses_project_name(self):
        d = dict(EMPTY_DRAFT)
        d["meta"] = dict(EMPTY_DRAFT["meta"], name="Riverview Station")
        resp = client.post("/export/excel", json=d)
        assert resp.status_code == 200
        cd = resp.headers.get("content-disposition", "")
        assert "Riverview" in cd or "Station" in cd

    def test_endpoint_returns_valid_xlsx_bytes(self):
        resp = client.post("/export/excel", json=full_draft())
        wb   = load_workbook(io.BytesIO(resp.content))
        assert set(EXPECTED_SHEETS).issubset(set(wb.sheetnames))

    def test_endpoint_empty_draft_200(self):
        resp = client.post("/export/excel", json={})
        assert resp.status_code == 200

    def test_endpoint_invalid_json_422(self):
        resp = client.post(
            "/export/excel",
            content=b"NOT JSON",
            headers={"Content-Type": "application/json"},
        )
        # FastAPI returns 422 Unprocessable Entity for invalid JSON body
        assert resp.status_code == 422

    def test_endpoint_partial_draft_hydraulics_only(self):
        d = dict(EMPTY_DRAFT)
        d["hydraulicsResult"] = HYDRAULICS_RESULT
        resp = client.post("/export/excel", json=d)
        assert resp.status_code == 200
        wb = load_workbook(io.BytesIO(resp.content))
        assert "System Curve" in wb.sheetnames
        assert "Pump Curves" in wb.sheetnames  # sheet exists but shows no-data note

    def test_endpoint_special_chars_in_project_name(self):
        d = dict(EMPTY_DRAFT)
        d["meta"] = dict(EMPTY_DRAFT["meta"], name="Station — A/B & C")
        resp = client.post("/export/excel", json=d)
        assert resp.status_code == 200
        cd = resp.headers.get("content-disposition", "")
        assert ".xlsx" in cd
